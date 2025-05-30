# excel_formatter.py
import os
import re
import logging
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Font, Side, Alignment, PatternFill
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from PIL import Image as PILImage
# from fpdf import FPDF
from alignment_utils import load_alignment_dict



from customer_utils import get_customer_logo  # for logo lookup


def sanitize_filename(filename):
    """
    Sanitize a string to be used as a valid filename or directory name.
    """
    filename = filename.strip()
    filename = re.sub(r'[<>:"/\\|?*]', '_', filename)
    return filename




def format_excel_file(xlsx_path):
    """
    Opens an XLSX file and applies two formatting steps:
      1. Auto‑adjusts each column’s width to the maximum width of its content.
      2. Applies a thin border around the header region (B1 to last column of row 5)
         and all cells in the dataframe portion (starting from row 6).
    """
    wb = load_workbook(xlsx_path)
    ws = wb.active

    # Turn off gridlines
    ws.sheet_view.showGridLines = False

    # Auto-adjust column widths for all columns.
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    # Header region formatting: from B1 to last column in row 5.
    min_col = 2  # Column B
    max_col = ws.max_column
    min_row = 1
    max_row = 5

    # # Insert "Prepared for" in the middle of the header region in row 1.
    # middle_col = (min_col + max_col) // 2
    # header_cell = ws.cell(row=1, column=middle_col, value="Prepared for")
    # header_cell.alignment = Alignment(horizontal="center")




    # Define the thin border style.
    thin_side = Side(border_style="thin", color="000000")

    # Format row 6: Bold and font size increase
    for cell in ws[6]:  # 6th row
        cell.font = Font(bold=True, size=12)  # Default size 10 → Increase by 2
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Create an outer border for the header region.
    for i in range(min_row, max_row + 1):
        for j in range(min_col, max_col + 1):
            cell = ws.cell(row=i, column=j)
            border_sides = {}
            if i == min_row:
                border_sides['top'] = thin_side
            if i == max_row:
                border_sides['bottom'] = thin_side
            if j == min_col:
                border_sides['left'] = thin_side
            if j == max_col:
                border_sides['right'] = thin_side
            if border_sides:
                cell.border = Border(**border_sides)

    # Apply thin border to cells starting from row 6.
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    for row in ws.iter_rows():
        if row[0].row >= 6:
            for cell in row:
                cell.border = border

    wb.save(xlsx_path)

def read_raw_data(xlsx_path):
    """
    Reads an Excel file and returns the raw dataframe data (skipping header details).
    """
    try:
        temp_df = pd.read_excel(xlsx_path, header=None)
        if (temp_df.shape[0] >= 5 and temp_df.shape[1] == 1) or \
           (temp_df.shape[0] >= 5 and isinstance(temp_df.iloc[0, 0], str) and
            ("Run Date:" in str(temp_df.iloc[1, 0]) or "Customer is" in str(temp_df.iloc[2, 0]))):
            raw_df = pd.read_excel(xlsx_path, skiprows=5)
        else:
            raw_df = pd.read_excel(xlsx_path)
        return raw_df
    except Exception as e:
        logging.error(f"Error reading raw data from {xlsx_path}: {e}")
        return None


def insert_logos_in_header(
    xlsx_path,
    left_logo_path,
    right_logo_path,
    target_width=151,
    target_height=76,
    left_padding=15,
    top_padding=15,
    right_padding=10,
    top_padding_right=5
):
    """
    Inserts logos into the header area of an XLSX file, if the image files exist.
    """
    wb = load_workbook(xlsx_path)
    ws = wb.active

    # ── LEFT LOGO ───────────────────────────────────────────────
    if left_logo_path and os.path.isfile(left_logo_path):
        try:
            # merge B1:C5
            ws.merge_cells("B1:C5")

            # load and scale
            img = OpenpyxlImage(left_logo_path)
            new_w, new_h = scale_to_fit(left_logo_path, target_width, target_height)
            img.width = new_w
            img.height = new_h

            # compute offset (in EMUs)
            off_x = (target_width - new_w) // 2 + left_padding
            off_y = (target_height - new_h) // 2 + top_padding
            marker = AnchorMarker(col=1, colOff=off_x * 9525, row=0, rowOff=off_y * 9525)
            size   = XDRPositiveSize2D(cx=new_w * 9525, cy=new_h * 9525)
            anchor = OneCellAnchor(_from=marker, ext=size)

            img.anchor = anchor
            ws.add_image(img)

        except Exception as e:
            logging.warning(f"Skipped left logo insertion: {e}")

    # ── RIGHT LOGO ──────────────────────────────────────────────
    max_col = ws.max_column
    if right_logo_path and os.path.isfile(right_logo_path) and max_col >= 2:
        try:
            # merge last two columns rows 1–5
            start = get_column_letter(max_col - 1)
            end   = get_column_letter(max_col)
            ws.merge_cells(f"{start}1:{end}5")

            # load and scale
            img = OpenpyxlImage(right_logo_path)
            new_w, new_h = scale_to_fit(right_logo_path, target_width, target_height)
            img.width  = new_w
            img.height = new_h

            # offset for right logo
            off_x = (target_width - new_w) // 2 + right_padding
            off_y = (target_height - new_h) // 2 + top_padding_right
            marker = AnchorMarker(col=(max_col - 2), colOff=off_x * 9525, row=0, rowOff=off_y * 9525)
            size   = XDRPositiveSize2D(cx=new_w * 9525, cy=new_h * 9525)
            anchor = OneCellAnchor(_from=marker, ext=size)

            img.anchor = anchor
            ws.add_image(img)

        except Exception as e:
            logging.warning(f"Skipped right logo insertion: {e}")

    # ── “Prepared for” LABEL ─────────────────────────────────────
    try:
        middle = max_col // 2
        cell = ws.cell(row=1, column=middle, value="Prepared for")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font      = Font(bold=True, size=14)
    except Exception:
        # fall back to next column if merged
        try:
            cell = ws.cell(row=1, column=middle + 1, value="Prepared for")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font      = Font(bold=True, size=14)
        except Exception as e:
            logging.warning(f"Could not insert 'Prepared for' label: {e}")

    wb.save(xlsx_path)



import os
import logging
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D


def ar_insert_logos_in_header(xlsx_path, left_logo_path, right_logo_path,
                              target_width=151, target_height=76,
                              left_padding=15, top_padding=15,
                              right_padding=10, top_padding_right=5):
    """
    Inserts logos into the header area of all sheets in an XLSX file for AR Aging Summary.
    """
    try:
        wb = load_workbook(xlsx_path)
        for ws in wb.worksheets:
            # Insert left logo in B1:C5
            try:
                ws.merge_cells("B1:C5")
                left_img = OpenpyxlImage(left_logo_path)
                new_w, new_h = scale_to_fit(left_logo_path, target_width, target_height)
                left_img.width, left_img.height = new_w, new_h
                left_img.anchor = 'B1'
                ws.add_image(left_img)
            except Exception as e:
                logging.error(f"Failed to insert left logo in sheet '{ws.title}': {e}")

            # Insert right logo in the last two columns
            max_col = ws.max_column
            if max_col >= 2:
                try:
                    last_two_start = get_column_letter(max_col - 1)
                    last_col_letter = get_column_letter(max_col)
                    merge_range = f"{last_two_start}1:{last_col_letter}5"
                    ws.merge_cells(merge_range)
                    right_img = OpenpyxlImage(right_logo_path)
                    new_w2, new_h2 = scale_to_fit(right_logo_path, target_width, target_height)
                    right_img.width, right_img.height = new_w2, new_h2
                    right_img.anchor = f"{last_col_letter}1"
                    ws.add_image(right_img)
                except Exception as e:
                    logging.error(f"Failed to insert right logo in sheet '{ws.title}': {e}")

        wb.save(xlsx_path)
        logging.info(f"Logos successfully inserted into all sheets of {xlsx_path}")

    except Exception as e:
        logging.error(f"Failed to process {xlsx_path} for AR logo insertion: {e}")





def add_report_details_to_xlsx(xlsx_path, report_name, customer_name, filter_column, customer_id, date_filter):
    try:
        df = pd.read_excel(xlsx_path)
    except Exception as e:
        logging.error(f"Error reading XLSX file {xlsx_path}: {e}")
        return

    run_date = datetime.now().strftime("%B %d, %Y at %I:%M:%S %p")

    today = datetime.now()
    bom = datetime(today.year, today.month, 1)
    eom = datetime(today.year, today.month + 1, 1) - pd.to_timedelta(1, unit='d')

    yesterday = today - pd.to_timedelta(1, unit='d')
    begining_of_last_month = datetime(today.year, today.month - 1, 1)
    end_of_last_month = datetime(today.year, today.month, 1) - pd.to_timedelta(1, unit='d')
    begining_of_this_week = today - pd.to_timedelta(today.weekday(), unit='d')
    end_of_this_week = today + pd.to_timedelta(6 - today.weekday(), unit='d')
    begining_of_last_week = today - pd.to_timedelta(today.weekday() + 7, unit='d')
    end_of_last_week = today - pd.to_timedelta(today.weekday() + 1, unit='d')
    last_12_months = today - pd.DateOffset(months=12)

    if date_filter == "Today":
        start_date = today
        end_date = today
    elif date_filter == "Yesterday":
        start_date = yesterday
        end_date = yesterday
    elif date_filter == "This Month":
        start_date = bom
        end_date = eom
    elif date_filter == "Last Month":
        start_date = begining_of_last_month
        end_date = end_of_last_month
    elif date_filter == "This Week":
        start_date = begining_of_this_week
        end_date = end_of_this_week
    elif date_filter == "Last Week":
        start_date = begining_of_last_week
        end_date = end_of_last_week
    elif date_filter == "Last 12 Months":
        start_date = last_12_months
        end_date = eom
    else:  # Default to today's date if none matches
        start_date = today
        end_date = today

    if report_name == "NEXTUS PATIENT BALANCE REPORT":
        report_period = f"{filter_column} is 365 Days"
    else:
        report_period = f"{filter_column} is between: {start_date.strftime('%m-%d-%Y')} - {end_date.strftime('%m-%d-%Y')}"

    # Removed PDF creation & handling code since we only need Excel output now.
    headers = [
        [report_name],
        [f"Run Date: {run_date}"],
        [f"Customer is {customer_name}"],
        [report_period],
        []
    ]

    final_output = headers + [df.columns.tolist()] + df.values.tolist()
    pd.DataFrame(final_output).to_excel(xlsx_path, index=False, header=False)

    # Apply yellow highlight to the Report Period cell.
    wb = load_workbook(xlsx_path)
    ws = wb.active
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    ws.cell(row=4, column=1).fill = yellow_fill
    wb.save(xlsx_path)

    format_excel_file(xlsx_path)

    # Logo handling remains unchanged.
    left_logo_path = os.path.join("static", "images", "Nextus-logo.png")
    right_logo_path = get_customer_logo(customer_id)
    if not os.path.exists(right_logo_path) or not right_logo_path:
        logging.warning(f"Right logo not found at {right_logo_path}, using fallback.")
        right_logo_path = os.path.join("static", "images", "no_name.png")

    insert_logos_in_header(xlsx_path, left_logo_path, right_logo_path,
                           target_width=151, target_height=76,
                           left_padding=15, top_padding=15,
                           right_padding=10, top_padding_right=5)

    logging.info(f"Report details and headers added correctly to {xlsx_path}")
    return report_period




# Fallback definitions for openpyxl drawing classes if needed.
try:
    from openpyxl.drawing.geometry import Ext
except ImportError:
    class Ext:
        def __init__(self, cx, cy):
            self.cx = cx
            self.cy = cy

try:
    from openpyxl.drawing.xdr import XDRPositiveSize2D
except ImportError:
    class XDRPositiveSize2D:
        def __init__(self, cx, cy):
            self.cx = cx
            self.cy = cy



import os
import logging
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


# from fpdf import FPDF




def ar_add_report_details_to_xlsx(xlsx_path, report_name, customer_name, filter_column, customer_id, date_filter):
    try:
        # Load the workbook (supports multiple sheets)
        wb = load_workbook(xlsx_path)
        all_sheets = wb.sheetnames
    except Exception as e:
        logging.error(f"Error reading XLSX file {xlsx_path}: {e}")
        return

    # Prepare Report Metadata
    run_date = datetime.now().strftime("%B %d, %Y at %I:%M:%S %p")
    today = datetime.now()
    bom = datetime(today.year, today.month, 1)
    eom = datetime(today.year, today.month + 1, 1) - pd.to_timedelta(1, unit='d')
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Date Filter Logic
    date_ranges = {
        "Today": (today, today),
        "Yesterday": (today - pd.to_timedelta(1, unit='d'), today - pd.to_timedelta(1, unit='d')),
        "This Month": (bom, eom),
        "Last Month": (datetime(today.year, today.month - 1, 1), bom - pd.to_timedelta(1, unit='d')),
        "This Week": (today - pd.to_timedelta(today.weekday(), unit='d'), today + pd.to_timedelta(6 - today.weekday(), unit='d')),
        "Last Week": (today - pd.to_timedelta(today.weekday() + 7, unit='d'), today - pd.to_timedelta(today.weekday() + 1, unit='d')),
        "Last 12 Months": (today - pd.DateOffset(months=12), eom)
    }
    start_date, end_date = date_ranges.get(date_filter, (today, today))

    # Add the report_period logic here

    report_period = f"{filter_column} is between: {start_date.strftime('%m-%d-%Y')} - {end_date.strftime('%m-%d-%Y')}"


    # Logo Paths
    left_logo_path = os.path.join("static", "images", "Nextus-logo.png")
    right_logo_path = get_customer_logo(customer_id)
    if not os.path.exists(right_logo_path):
        right_logo_path = os.path.join("static", "images", "no_name.png")

    # Process each sheet
    for sheet_name in all_sheets:
        ws = wb[sheet_name]

        # Insert Report Metadata
        ws.insert_rows(1, 5)
        ws["A1"] = report_name
        ws["A2"] = f"Run Date: {run_date}"
        ws["A3"] = f"Customer is {customer_name}"
        ws["A4"] = report_period
        ws["A4"].fill = yellow_fill

        # Insert logos for each sheet
        ar_insert_logos_in_header(xlsx_path, left_logo_path, right_logo_path,
                               target_width=151, target_height=76,
                               left_padding=15, top_padding=15,
                               right_padding=10, top_padding_right=5)

    # Save Transformed Excel
    wb.save(xlsx_path)
    logging.info(f"Report details and logos added for all sheets in {xlsx_path}")

    return report_period



def consolidate_monthly_report(customer_folder, report_name, current_year, current_month):
    """
    Consolidates customer reports for a specific report and month into a single report.
    """
    import os
    logging.info(f"Consolidating reports in {customer_folder}")
    all_customer_data = []
    report_month_folder = os.path.join(customer_folder, current_year, current_month, report_name)
    try:
        customer_folders = [os.path.join(report_month_folder, customer)
                            for customer in os.listdir(report_month_folder)
                            if os.path.isdir(os.path.join(report_month_folder, customer))]
    except Exception as e:
        logging.error(f"Error listing customer folders in {report_month_folder}: {e}")
        return None

    for cust_folder in customer_folders:
        consolidated_file_path = os.path.join(cust_folder, f"{os.path.basename(cust_folder)}_{current_month}_consolidated.xlsx")
        if os.path.exists(consolidated_file_path):
            customer_data = read_raw_data(consolidated_file_path)
            if customer_data is not None:
                all_customer_data.append(customer_data)
            else:
                logging.error(f"Error reading raw data from {consolidated_file_path}. Skipping.")
        else:
            logging.info(f"Consolidated file not found for {cust_folder}. Skipping.")

    if all_customer_data:
        monthly_consolidated_data = pd.concat(all_customer_data, ignore_index=True)
        monthly_consolidated_data.drop_duplicates(inplace=True)
        consolidated_report_path = os.path.join(report_month_folder, f"{report_name}_{current_month}_consolidated.xlsx")
        monthly_consolidated_data.to_excel(consolidated_report_path, index=False)
        format_excel_file(consolidated_report_path)
        logging.info(f"Monthly consolidated report saved at {consolidated_report_path}")
        return consolidated_report_path
    else:
        logging.info("No data to consolidate for this month.")
        return None

def update_consolidated_file(customer_folder, new_file_path, customer_name):
    """
    Updates the consolidated XLSX file by appending new data, removing duplicates, and formatting.
    """
    current_month = datetime.now().strftime("%b")
    consolidated_file_name = f"{customer_name}_{current_month}_consolidated.xlsx"
    consolidated_file_path = os.path.join(customer_folder, consolidated_file_name)

    try:
        new_data = pd.read_excel(new_file_path, skiprows=5)
    except Exception as e:
        logging.error(f"Error reading new data from {new_file_path}: {e}")
        return

    if os.path.exists(consolidated_file_path):
        consolidated_data = read_raw_data(consolidated_file_path)
        if consolidated_data is None:
            consolidated_data = pd.DataFrame()
        updated_data = pd.concat([consolidated_data, new_data]).drop_duplicates()
        updated_data.to_excel(consolidated_file_path, index=False)
    else:
        new_data.to_excel(consolidated_file_path, index=False)

    format_excel_file(consolidated_file_path)
    logging.info(f"Consolidated file updated at {consolidated_file_path}")

def scale_to_fit(image_path, target_width, target_height):
    """
    Scales an image to fit within the target dimensions while preserving aspect ratio.
    """
    with PILImage.open(image_path) as img:
        orig_w, orig_h = img.size
    scale = min(target_width / orig_w, target_height / orig_h)
    new_w = int(orig_w * scale)
    new_h = int(orig_h * scale)
    return new_w, new_h


# Function for dynamic alignment
def dynamic_column_alignment(xlsx_path, report_name):
    """
    Aligns columns dynamically for specific reports using alignment rules from the JSON file.
    Applies alignment directly in the Excel file.
    """
    alignment_data = load_alignment_dict()
    wb = load_workbook(xlsx_path)
    ws = wb.active
    max_row = ws.max_row
    max_col = ws.max_column

    # Get the alignment rules for this report (keys must match report_name exactly)
    report_alignment = alignment_data.get(report_name, {})

    # Mapping for openpyxl
    align_map = {"L": "left", "C": "center", "R": "right"}

    # Loop through data cells (from row 6)
    for row in range(6, max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            col_rule = report_alignment.get(str(col), {"align": "L"})
            horizontal_align = align_map.get(col_rule.get("align", "L").upper(), "left")
            cell.alignment = Alignment(horizontal=horizontal_align)
    
    wb.save(xlsx_path)
    logging.info(f"Dynamic alignment applied for {report_name}.")

def process_monthly_consolidation(base_output_folder, report_name, current_year, current_month):
    """
    Processes the consolidation of all customer reports for a specific report and month.
    """
    report_month_folder = os.path.join(base_output_folder, current_year, current_month, report_name)
    logging.info(f"Attempting to create folder: {report_month_folder}")
    try:
        os.makedirs(report_month_folder, exist_ok=True)
        logging.info(f"Folder created successfully or already exists: {report_month_folder}")
    except Exception as e:
        logging.error(f"Error creating folder {report_month_folder}: {e}")

    consolidated_report_path = consolidate_monthly_report(base_output_folder, report_name, current_year, current_month)
    if consolidated_report_path:
        logging.info(f"Consolidated report generated at: {consolidated_report_path}")
    else:
        logging.info(f"No consolidated report generated for {report_name} in {current_month}.")



# alignment_dict = {
#     "Activity Summary by Month Report": {1: {"align": "L", "width": 25}, 2: {"align": "C", "width": 25}, 3: {"align": "C", "width": 30}, 4: {"align": "C", "width": 50}, 5: {"align": "C", "width": 50}},
#     "CLAIMS BILLED REPORT": {1: {"align": "L", "width": 23}, 2: {"align": "C", "width": 40}, 3: {"align": "C", "width": 32}, 4: {"align": "C", "width": 30}, 5: {"align": "C", "width": 32}, 6: {"align": "C", "width": 40}, 7: {"align": "C", "width": 30}, 8: {"align": "C", "width": 30}, 9: {"align": "L", "width": 41}, 10: {"align": "L", "width": 95}, 11: {"align": "R", "width": 82}},    
#     "NEXTUS MONTHLY DETAIL PAYMENT REPORT": {1: {"align": "L", "width": 60}, 2: {"align": "C", "width": 27}, 3: {"align": "C", "width": 28}, 4: {"align": "L", "width": 50}, 5: {"align": "L", "width": 25}, 6: {"align": "C", "width": 30}, 7: {"align": "C", "width": 20}, 8: {"align": "C", "width": 29}, 9: {"align": "C", "width": 27}, 10: {"align": "C", "width": 30}, 11: {"align": "C", "width": 30}, 12: {"align": "L", "width": 75}, 13: {"align": "R", "width": 25}, 14: {"align": "C", "width": 19}},
#     "NEXTUS PAYMENT SUMMARY MONTHLY REPORT": {1: {"align": "L", "width": 127}, 2: {"align": "C", "width": 42}, 3: {"align": "C", "width": 45}, 4: {"align": "C", "width": 42}, 5: {"align": "L", "width": 127}, 6: {"align": "C", "width": 37}, 7: {"align": "R", "width": 38}},
#     "NEXTUS PAYMENT SUMMARY WEEKLY REPORT": {1: {"align": "L", "width": 127}, 2: {"align": "C", "width": 42}, 3: {"align": "C", "width": 45}, 4: {"align": "C", "width": 37}, 5: {"align": "L", "width": 135}, 6: {"align": "C", "width": 37}, 7: {"align": "R", "width": 38}},
#     "AR AGING SUMMARY": {1: {"align": "L", "width": 135}, 2: {"align": "C", "width": 63}, 3: {"align": "C", "width": 49}, 4: {"align": "C", "width": 42}, 5: {"align": "C", "width": 42}, 6: {"align": "C", "width": 70}, 7: {"align": "C", "width": 43}, 8: {"align": "C", "width": 45}, 9: {"align": "C", "width": 38}, 10: {"align": "C", "width": 35}, 11: {"align": "C", "width": 35}},
#     "NEXTUS PATIENT BALANCE REPORT": {1: {"align": "L", "width": 17}, 2: {"align": "C", "width": 65}, 3: {"align": "C", "width": 35}, 4: {"align": "L", "width": 27}, 5: {"align": "L", "width": 35}, 6: {"align": "C", "width": 35}, 7: {"align": "C", "width": 35}, 8: {"align": "C", "width": 35}, 9: {"align": "C", "width": 45}, 10: {"align": "C", "width": 35}, 11: {"align": "C", "width": 35}, 12: {"align": "L", "width": 38}, 13: {"align": "R", "width": 35}
# }}