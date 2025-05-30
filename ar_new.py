import os
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, Font, Alignment
import re

import requests
import zipfile
import base64
import io
import pandas as pd
import os
import logging
from datetime import datetime
from auth import get_access_token
#connections
from excel_formatter import sanitize_filename



def process_ar_aging_summary(customer_id, customer_name, report_identifier, filter_identifier):
    """
    Handles AR Aging Summary report: Fetches, extracts, converts, and stores it.
    """
    identifier = get_report_identifier(customer_id, report_identifier, filter_identifier)
    if not identifier:
        logging.error(f"❌ Failed to get report identifier for {customer_name}.")
        return None

    binary_data = poll_for_report(customer_id, identifier)
    if not binary_data:
        logging.error(f"❌ Report did not complete successfully for {customer_name}.")
        return None

    # Decode ZIP file
    decoded_data = base64.b64decode(binary_data)
    zip_file = zipfile.ZipFile(io.BytesIO(decoded_data))

    # Extract CSV file
    extracted_csv_path = None
    for file_name in zip_file.namelist():
        if file_name.endswith(".csv"):
            extracted_csv_path = os.path.join(INPUT_FOLDER, f"{customer_name}_AR_DataSource.csv")
            zip_file.extract(file_name, INPUT_FOLDER)
            os.rename(os.path.join(INPUT_FOLDER, file_name), extracted_csv_path)
            break

    if not extracted_csv_path:
        logging.error(f"❌ No CSV file found in ZIP for {customer_name}.")
        return None

    # Convert CSV to Excel
    df = pd.read_csv(extracted_csv_path, encoding="utf-8")
    excel_file_path = extracted_csv_path.replace(".csv", ".xlsx")
    df.to_excel(excel_file_path, index=False)

    # Cleanup CSV file
    os.remove(extracted_csv_path)
    logging.info(f"✅ AR Summary DataSource saved: {excel_file_path}")

    return excel_file_path  # Returning the path for next steps


# def sanitize_filename(filename):
#     """
#     Sanitize a string to be used as a valid filename or directory name.
#     """
#     filename = filename.strip()
#     filename = re.sub(r'[<>:"/\\|?*]', '_', filename)
#     return filename

# -----------------------------------------------
# 1) Generate Pivot DataFrames & Determine Customer
# -----------------------------------------------
def generate_pivots_and_customer(input_file: str):
    """
    Reads 'Test Claims Report' from 'input_file',
    extracts 'Customer Name' from the DataFrame (assuming it's consistent),
    and creates 4 pivot DataFrames.
    Returns (df_ar_aging, df_payer_type, df_current_payer, df_priority, customer_name).
    """

    # Read the raw data (which now includes "Customer ID" and "Customer Name")
    df = pd.read_excel(input_file)

    # Ensure charge balance values are numeric (strip $ and commas)
    monetary_cols = [
        "Charge Balance", "Charge Balance Due Ins", "Charge Balance Due Pat",
        "Charge Balance Due Other", "Charge Balance At Collections"
    ]

    for col in monetary_cols:
        df[col] = df[col].replace('[\$,]', '', regex=True).astype(float)

    # df = pd.read_excel(input_file, sheet_name="Test Claims Report")

    # Determine the customer name from the first non-null row
    # (Adjust if your data can have multiple customers in one file)
    customer_name = df["Customer Name"].dropna().unique()[0]

    # 1. A/R Aging Summary
    df_ar_aging_summary = pd.pivot_table(
        df,
        index="Charge Fromdate Age",
        values=[
            "Charge Balance",
            "Charge Balance Due Ins",
            "Charge Balance Due Pat",
            "Charge Balance Due Other",
            "Charge Balance At Collections"
        ],
        aggfunc="sum",
        margins=True,
        margins_name=""
    ).reset_index().fillna(0)

    # Force the column order
    desired_order = [
        "Charge Fromdate Age",
        "Charge Balance",
        "Charge Balance Due Ins",
        "Charge Balance Due Pat",
        "Charge Balance Due Other",
        "Charge Balance At Collections"
    ]
    df_ar_aging_summary = df_ar_aging_summary[desired_order]

    # 2. Payer Type Breakdown
    df_payer_type_breakdown = pd.pivot_table(
        df[df["Charge Balance Due"] == "Insurance"],
        index="Charge Current Payer Type",
        columns="Charge Fromdate Age",
        values="Charge Balance",
        aggfunc="sum",
        margins=True,
        margins_name=""
    ).fillna(0).reset_index()

    # 3. Current Payer Breakdown
    df_current_payer_breakdown = pd.pivot_table(
        df[df["Charge Balance Due"] == "Insurance"],
        index="Charge Current Payer Name",
        columns="Charge Fromdate Age",
        values="Charge Balance",
        aggfunc="sum",
        margins=True,
        margins_name=""
    ).fillna(0).reset_index()

    # 4. Total Balance by Payer Priority
    df_total_balance_by_payer_priority = pd.pivot_table(
        df,
        index=[],  # single summary row
        columns="Charge Current Payer Priority",
        values="Charge Balance",
        aggfunc="sum",
        margins=True,
        margins_name="Grand Total"
    ).fillna(0)

    # Remove built-in 'Grand Total' if present, add (Primary+Secondary+Tertiary)
    if "Grand Total" in df_total_balance_by_payer_priority.columns:
        df_total_balance_by_payer_priority.drop("Grand Total", axis=1, inplace=True)

    desired_cols = ["Primary", "Secondary", "Tertiary"]
    existing = [c for c in desired_cols if c in df_total_balance_by_payer_priority.columns]
    df_total_balance_by_payer_priority["Grand Total"] = df_total_balance_by_payer_priority[existing].sum(axis=1)

    rename_map = {
        "Primary": "Charge Primary Balance (Sum)",
        "Secondary": "Charge Secondary Balance (Sum)",
        "Tertiary": "Charge Tertiary Balance (Sum)",
        "Grand Total": "Charge Grand Total (Sum)",
    }
    df_total_balance_by_payer_priority = df_total_balance_by_payer_priority.rename(columns=rename_map)

    return (
        df_ar_aging_summary,
        df_payer_type_breakdown,
        df_current_payer_breakdown,
        df_total_balance_by_payer_priority,
        customer_name
    )

# -----------------------------------------------
# 2) Write All Sheets to One Workbook
# -----------------------------------------------
def write_all_sheets(
    output_file: str,
    df_ar_aging_summary: pd.DataFrame,
    df_payer_type_breakdown: pd.DataFrame,
    df_current_payer_breakdown: pd.DataFrame,
    df_total_balance_by_payer_priority: pd.DataFrame
):
    """
    Writes the 4 pivot DataFrames plus the entire input data (sheet "Input Data")
    into one Excel workbook. The pivot data starts at row=5 so we have room
    for custom headers in rows 1-3.
    """
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        df_ar_aging_summary.to_excel(
            # writer, sheet_name="AR Aging Summary", index=False, startrow=5
            writer, sheet_name = "AR Aging Summary", index = False, startrow =1
        )
        df_payer_type_breakdown.to_excel(
            writer, sheet_name="Payer Type Breakdown", index=False, startrow=1
        )
        df_current_payer_breakdown.to_excel(
            writer, sheet_name="Current Payer Breakdown", index=False, startrow=1
        )
        df_total_balance_by_payer_priority.to_excel(
            writer, sheet_name="Total Balance by Payer Priority", index=False, startrow=1
        )


# -----------------------------------------------
# 3) Apply Formatting (headers, run date, currency, etc.)
# -----------------------------------------------
def format_all_pivots(
    output_file: str,
    customer_name: str,
    df_ar_aging_summary: pd.DataFrame,
    df_payer_type_breakdown: pd.DataFrame,
    df_current_payer_breakdown: pd.DataFrame,
    df_total_balance_by_payer_priority: pd.DataFrame
):
    """
    Opens 'output_file' with openpyxl, applies formatting to the pivot sheets,
    then saves it. The pivot data is assumed to start at row=5, so row=5 is the header,
    row=6+ is data. Rows 1-3 are used for pivot name, run date, and customer info.
    """
    wb = load_workbook(output_file)
    run_dt = datetime.now()  # Current date/time

    # List of pivot sheets to format
    pivot_sheets = [
        ("AR Aging Summary", df_ar_aging_summary),
        ("Payer Type Breakdown", df_payer_type_breakdown),
        ("Current Payer Breakdown", df_current_payer_breakdown),
        ("Total Balance by Payer Priority", df_total_balance_by_payer_priority),
    ]

    for sheet_name, df_pivot in pivot_sheets:
        ws = wb[sheet_name]
        format_pivot_sheet(
            ws,
            pivot_name=sheet_name,
            customer_name=customer_name,
            run_datetime=run_dt,
            df=df_pivot,
            # start_row=6
            start_row=1
        )

    wb.save(output_file)
    print(f"Formatted pivot file saved at: {output_file}")

def format_pivot_sheet(
    ws,
    pivot_name: str,
    customer_name: str,
    run_datetime: datetime,
    df: pd.DataFrame,
    # start_row: int = 6
    start_row: int = 1
):
    """
    - A1: Pivot Name in size=16, bold
    - A2: "Run Date: {Month DD, YYYY at HH:MM:SS AM/PM}"
    - A3: "Customer is {customer_name}"
    - Row=4: blank
    - Row=5: pivot header
    - Row=6+: pivot data
    - Auto-fit columns, thin border, $#,##0.00 for numeric columns
    """

    header_row = start_row
    max_row = ws.max_row
    max_col = ws.max_column

    # Merge A1 to the last column (e.g., A1:Z1)
    #Eliminate merging
    # last_col_letter = get_column_letter(max_col)
    # ws.merge_cells(f"A1:{last_col_letter}1")
    # ws.merge_cells(f"A2:{last_col_letter}2")
    # ws.merge_cells(f"A3:{last_col_letter}3")
    # ws.merge_cells(f"A4:{last_col_letter}4")

    # # 1) Fill row 1..4
    # ws["A1"] = pivot_name
    # ws["A1"].font = Font(size=16, bold=True)
    # ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    #
    # run_date_str = run_datetime.strftime("%B %d, %Y at %I:%M:%S %p")
    # ws["A2"] = f"Run Date: {run_date_str}"
    # ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    #
    # ws["A3"] = f"Customer is {customer_name}"
    # ws["A3"].alignment = Alignment(horizontal="left", vertical="center")
    #
    # ws["A4"] = f"Charge From Date is between "
    # ws["A4"].alignment = Alignment(horizontal="left", vertical="center")



    # 2) Apply formatting from row 5 downward
    max_row = ws.max_row
    max_col = ws.max_column

    # # Auto-fit columns
    # for col_idx in range(1, max_col + 1):
    #     col_letter = get_column_letter(col_idx)
    #     # Header is at row `start_row`
    #     header_val = ws.cell(row=start_row, column=col_idx).value
    #     max_len = len(str(header_val)) if header_val else 0
    #     for row_idx in range(start_row + 1, max_row + 1):
    #         cell_val = ws.cell(row=row_idx, column=col_idx).value
    #         if cell_val is not None:
    #             max_len = max(max_len, len(str(cell_val)))
    #     ws.column_dimensions[col_letter].width = max_len + 2

    for col_idx in range(1, max_col + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0

        # Scan ALL rows from 1 to max_row
        for row_idx in range(1, max_row + 1):
            cell_val = ws.cell(row=row_idx, column=col_idx).value

            if cell_val is not None:
                # *** ADDED *** if numeric, format with commas before measuring length
                if isinstance(cell_val, (int, float)):
                    # e.g., 4513436.08 -> "4,513,436.08"
                    val_str = f"{cell_val:,.2f}"
                else:
                    val_str = str(cell_val)

                max_len = max(max_len, len(val_str))


        ws.column_dimensions[col_letter].width = max_len + 4

    # Thin border
    thin_side = Side(border_style="thin", color="000000")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    # Define border styles
    thin_side = Side(border_style="thin", color="000000")
    double_side = Side(border_style="double", color="000000")

    # Apply a single (thin) top border to the header row
    for col_idx in range(1, max_col + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.border = Border(top=thin_side)



    # Apply a double bottom border to the last row
    for col_idx in range(1, max_col + 1):
        cell = ws.cell(row=max_row-1, column=col_idx)
        cell.border = Border(bottom=double_side)
        cell = ws.cell(row=max_row, column=col_idx)
        cell.font = Font(size=12, bold=True)

    # Left-align & set font=12, bold for header row
    # -------------------------------------------------------------
    for col_idx in range(1, max_col + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.alignment = Alignment(horizontal="left")
        cell.font = Font(size=12, bold=True)


    # Currency format for all numeric columns in df
    numeric_cols = df.select_dtypes(include='number').columns.tolist()

    # Map col_name -> col_index
    col_map = {}
    for i, col_name in enumerate(df.columns, start=1):
        col_map[col_name] = i

    for col_name in numeric_cols:
        if col_name in col_map:
            col_idx = col_map[col_name]
            for row_idx in range(start_row + 1, max_row + 1):
                ws.cell(row=row_idx, column=col_idx).number_format = '$#,##0.00'


# -----------------------------------------------
# Combined function to process the file
# -----------------------------------------------


# def process_claim_file(input_file_path: str, output_file_path: str):
#     """
#     Processes the claim file:
#       1) Generates pivot DataFrames and extracts the customer name.
#       2) Writes the pivots to an Excel workbook.
#       3) Applies formatting to each pivot sheet.
#     """
#     (
#         df_ar_aging_summary,
#         df_payer_type_breakdown,
#         df_current_payer_breakdown,
#         df_total_balance_by_payer_priority,
#         customer_name
#     ) = generate_pivots_and_customer(input_file_path)
#
#     # Extract directory from the provided output file path.
#     output_dir = os.path.dirname(output_file_path)
#
#     # Create output file path with customer name included.
#     # output_file_path = os.path.join(output_dir, f"A_R Aging Summary - {customer_name}.xlsx")
#
#     # sanitized_customer_name = sanitize_filename(customer_name)
#     # output_file_path = os.path.join(output_dir, f"A_R Aging Summary - {sanitized_customer_name}.xlsx")
#
#     # sanitized_customer_name = sanitize_filename(customer_name)  # Remove spaces
#     sanitized_customer_name = sanitize_filename(customer_name).replace(" ", "")  # Remove spaces
#     # output_file_path = os.path.join(output_dir, f"A_R Aging Summary - {sanitized_customer_name}.xlsx")
#     output_file_path = os.path.join(customer_folder, new_file_name)
#
#
#     # Write all pivot DataFrames to the output workbook.
#     write_all_sheets(
#         output_file=output_file_path,
#         df_ar_aging_summary=df_ar_aging_summary,
#         df_payer_type_breakdown=df_payer_type_breakdown,
#         df_current_payer_breakdown=df_current_payer_breakdown,
#         df_total_balance_by_payer_priority=df_total_balance_by_payer_priority
#     )
#
#     # Apply formatting to all pivot sheets.
#     format_all_pivots(
#         output_file=output_file_path,
#         customer_name=customer_name,
#         df_ar_aging_summary=df_ar_aging_summary,
#         df_payer_type_breakdown=df_payer_type_breakdown,
#         df_current_payer_breakdown=df_current_payer_breakdown,
#         df_total_balance_by_payer_priority=df_total_balance_by_payer_priority
#     )
#
#     print(f"Running process_claim_file with: input={input_file_path}, output={output_file_path}")

def process_claim_file(input_file_path: str, output_file_path: str):
    """
    Processes the claim file:
      1) Generates pivot DataFrames and extracts the customer name.
      2) Writes the pivots to an Excel workbook.
      3) Applies formatting to each pivot sheet.
    """
    # Generate pivots and extract customer name
    (
        df_ar_aging_summary,
        df_payer_type_breakdown,
        df_current_payer_breakdown,
        df_total_balance_by_payer_priority,
        customer_name
    ) = generate_pivots_and_customer(input_file_path)

    # Ensure output folder exists
    os.makedirs(os.path.dirname(output_file_path), exist_ok=True)

    # Write all pivot DataFrames to the output workbook.
    write_all_sheets(
        output_file=output_file_path,
        df_ar_aging_summary=df_ar_aging_summary,
        df_payer_type_breakdown=df_payer_type_breakdown,
        df_current_payer_breakdown=df_current_payer_breakdown,
        df_total_balance_by_payer_priority=df_total_balance_by_payer_priority
    )

    # Apply formatting to all pivot sheets.
    format_all_pivots(
        output_file=output_file_path,
        customer_name=customer_name,
        df_ar_aging_summary=df_ar_aging_summary,
        df_payer_type_breakdown=df_payer_type_breakdown,
        df_current_payer_breakdown=df_current_payer_breakdown,
        df_total_balance_by_payer_priority=df_total_balance_by_payer_priority
    )

    print(f"✅ process_claim_file completed: input={input_file_path}, output={output_file_path}")





# def process_claim_file(input_file_path: str, output_file_path: str):
#     """
#     Processes the claim file:
#       1) Generates pivot DataFrames and extracts the customer name.
#       2) Writes the pivots to an Excel workbook.
#       3) Applies formatting to each pivot sheet.
#     """
#     (
#         df_ar_aging_summary,
#         df_payer_type_breakdown,
#         df_current_payer_breakdown,
#         df_total_balance_by_payer_priority,
#         customer_name
#     ) = generate_pivots_and_customer(input_file_path)
#
#     # Extract directory from the provided output file path.
#     output_dir = os.path.dirname(output_file_path)
#
#     # Create output file path with customer name included.
#     sanitized_customer_name = sanitize_filename(customer_name).replace(" ", "")  # Remove spaces
#     output_file_path = os.path.join(output_dir, f"A_R Aging Summary - {sanitized_customer_name}.xlsx")
#
#     # Write all pivot DataFrames to the output workbook.
#     write_all_sheets(
#         output_file=output_file_path,
#         df_ar_aging_summary=df_ar_aging_summary,
#         df_payer_type_breakdown=df_payer_type_breakdown,
#         df_current_payer_breakdown=df_current_payer_breakdown,
#         df_total_balance_by_payer_priority=df_total_balance_by_payer_priority
#     )
#
#     # Apply formatting to all pivot sheets.
#     format_all_pivots(
#         output_file=output_file_path,
#         customer_name=customer_name,
#         df_ar_aging_summary=df_ar_aging_summary,
#         df_payer_type_breakdown=df_payer_type_breakdown,
#         df_current_payer_breakdown=df_current_payer_breakdown,
#         df_total_balance_by_payer_priority=df_total_balance_by_payer_priority
#     )
#
#     print(f"Running process_claim_file with: input={input_file_path}, output={output_file_path}")








# from openpyxl import load_workbook
# from openpyxl.styles import Font, Alignment, PatternFill
# import logging
# def format_all_pivot_sheets(xlsx_path, report_name, customer_name, filter_column, customer_id, date_filter):
#     """
#     Formats all pivot sheets in an Excel workbook and applies report details to each sheet.
#     Ensures proper column width, alignment, header formatting, and inserts report details.
#     """
#     try:
#         # Load the workbook
#         wb = load_workbook(xlsx_path)
#
#         for sheet_name in wb.sheetnames:
#             ws = wb[sheet_name]
#
#             # Apply Formatting to Headers (First Row)
#             for col in ws.iter_cols(min_row=1, max_row=1):
#                 for cell in col:
#                     cell.font = Font(bold=True)  # Bold Headers
#                     cell.alignment = Alignment(horizontal="center")  # Center Alignment
#
#             # Auto-adjust Column Widths
#             for col in ws.columns:
#                 max_length = 0
#                 col_letter = col[0].column_letter  # Get column letter
#                 for cell in col:
#                     try:
#                         if cell.value:
#                             max_length = max(max_length, len(str(cell.value)))
#                     except:
#                         pass
#                 ws.column_dimensions[col_letter].width = max_length + 2  # Add padding
#
#         # Save formatted workbook
#         wb.save(xlsx_path)
#
#         # Call add_report_details_to_xlsx for every pivot sheet
#         report_period = add_report_details_to_xlsx(
#             xlsx_path, report_name, customer_name, filter_column, customer_id, date_filter
#         )
#
#         print(f"✅ Formatting and report details applied to all pivot sheets in: {xlsx_path}")
#         return report_period
#
#     except Exception as e:
#         logging.error(f"❌ ERROR formatting pivot sheets in {xlsx_path}: {e}")
#         return None

# # Example usage:
# if __name__ == "__main__":
#     input_file = "input/MILTON AR Summary Datasource.xlsx"
#     output_file = "./output/"  # Output file path will be generated based on customer name""
#     process_claim_file(input_file, output_file)
